# #For Process RN657545267IN
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#

from xlsxwriter import Workbook


from PIL import Image, ImageDraw, ImageFont

from io import BytesIO
import sys
import zipfile

from base64 import b64decode
import pandas as pd
import requests
import os
from time import sleep,time
import datetime
import base64

import barcode
from barcode.writer import ImageWriter

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.cell.cell import WriteOnlyCell

from tkinter.filedialog import askdirectory



#For Status Extraction
def start(df,i,l,sleep_,pdf_opt):
    chrome_options = Options()
#     chrome_options.add_argument('--headless')
#     chrome_options.add_argument('--disable-gpu')
#     chrome_options.add_argument('--no-sandbox')
        
    driver = webdriver.Chrome(options=chrome_options)
    driver.get('https://app.indiapost.gov.in/customer-selfservice/login')
    driver.find_element(By.XPATH,"//button[@class='flex-1 py-3 px-3 rounded-full text-sm font-medium transition-all duration-500 ease-in-out text-red-700 bg-transparent']").click()
    driver.find_element(By.XPATH,"//input[@class='bg-white px-4 py-2.5 text-sm focus:outline-none w-full rounded-lg [appearance:textfield] [&::-webkit-outer-spin-button]:appearance-none [&::-webkit-inner-spin-button]:appearance-none rounded-s-none']").send_keys('9490017975')
    driver.find_element(By.XPATH,"//button[@class='items-center gap-2 whitespace-nowrap text-sm font-medium ring-offset-background duration-300 focus-visible:outline-none focus-visible:ring-2 focus-visible:ring-ring focus-visible:ring-offset-2 disabled:pointer-events-none disabled:opacity-50 [&_svg]:pointer-events-none [&_svg]:size-4 [&_svg]:shrink-0 relative overflow-hidden flex justify-center mx-auto bg-red-700 text-white hover:bg-red-800 transition-colors py-3 h-11 rounded-md px-8 w-full mt-8']").click()
    otp = input('Enter otp :- ')
    driver.find_element(By.XPATH,"//input[@class='bg-white px-4 py-2.5 text-sm focus:outline-none rounded-lg [appearance:textfield] [&::-webkit-outer-spin-button]:appearance-none [&::-webkit-inner-spin-button]:appearance-none w-12 h-12 text-center']").send_keys(otp)
    sleep(2)
    driver.find_element(By.XPATH,"//button[@class='items-center gap-2 whitespace-nowrap text-sm font-medium ring-offset-background duration-300 focus-visible:outline-none focus-visible:ring-2 focus-visible:ring-ring focus-visible:ring-offset-2 disabled:pointer-events-none disabled:opacity-50 [&_svg]:pointer-events-none [&_svg]:size-4 [&_svg]:shrink-0 relative overflow-hidden flex justify-center mx-auto bg-red-700 text-white hover:bg-red-800 transition-colors py-3 h-11 rounded-md px-8 w-full mt-8']").click()
    sleep(5)
    driver.get('https://app.indiapost.gov.in/customer-selfservice/bulk-articles-tracking')
        
    pdfs = []
    
    df = df[i:l+1]
    df_view = st.empty()
    df_view.dataframe(df)
    cols = st.columns(4)
    with st.spinner('Please wait..'):
        sleep(1)
    with st.status("Processing.....",expanded=True):
        ot = time()
        rt = 0
        c = 0
        wait = WebDriverWait(driver, 10)
        wait_ = WebDriverWait(driver, sleep_)
        # Assuming 'i', 'l', 'j', 'rt' are initialized correctly before the loop
        j = 0
        while i <= l:
            # --- START: LOGIC REWORK (Part 1: Building the batch) ---
            
            # This check using a stale 'j' is removed as it's unreliable.
            # The check is now done properly inside the 'for' loop.
            
            ref = ''
            items_in_this_batch = 0 # Use a clear counter instead of manipulating 'j'
            
            # This loop correctly builds the batch and finds its size
            for batch_offset in range(10):
                current_index = i + batch_offset
                if current_index > l:
                    break # We've hit the end of the dataframe
                
                ref += df.loc[current_index, 'RPAD Barcode No '] + ','
                items_in_this_batch += 1

            # If the last batch was full and there are no more items, stop.
            if items_in_this_batch == 0:
                break
                
            # --- END: LOGIC REWORK (Part 1) ---


            # --- MANDATORY SELENIUM SCRIPT (UNCHANGED AS REQUESTED) ---
            if rt == 0:
                rt = time()
            
            ip = wait.until(EC.presence_of_element_located((By.XPATH, "//textarea[@class='flex w-full rounded-md border border-input bg-background px-3 py-2 text-base ring-offset-background placeholder:text-muted-foreground focus-visible:outline-none focus-visible:ring-2 focus-visible:ring-ring focus-visible:ring-offset-2 disabled:cursor-not-allowed disabled:opacity-50 md:text-sm min-h-32']")))
            ip.clear()
            ip.send_keys(ref)
            driver.find_element(By.XPATH, "//button[@class='gap-2 whitespace-nowrap text-sm ring-offset-background duration-300 focus-visible:outline-none focus-visible:ring-2 focus-visible:ring-ring focus-visible:ring-offset-2 disabled:pointer-events-none [&_svg]:pointer-events-none [&_svg]:size-4 [&_svg]:shrink-0 relative overflow-hidden w-11/12 mx-auto h-10 min-w-[8rem] bg-red-600 hover:bg-red-700 text-white font-medium rounded-md py-2 px-4 flex items-center justify-center transition-colors disabled:opacity-50 disabled:cursor-not-allowed']").click()
            sleep(4)
            
            try:
                wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//h3[@class='text-lg font-semibold text-gray-800']")
                    )
                )
            except Exception:
                # We must still advance 'i' to avoid getting stuck in an infinite loop
                i += items_in_this_batch
                continue

            driver.find_element(By.XPATH,"//div[@class='flex gap-2 items-center bg-orange-100 justify-center min-w-[120px] px-3 text-orange-500 rounded-lg border-2 border-orange-500 p-1']").click()
            table = driver.find_element(By.XPATH,"//table[@class='w-full caption-bottom text-sm min-w-full']")
            details = table.find_elements(By.XPATH,"//tbody//tr[@class='border-b transition-colors hover:bg-gray-100 data-[state=selected]:bg-gray-100 odd:bg-white even:bg-gray-50']")
            # --- END OF MANDATORY SELENIUM SCRIPT ---


            # --- START: LOGIC REWORK (Part 2: Processing the results) ---
            data = []
            print(len(details), 'this is length')

            # THE CRITICAL FIX: Loop over the number of results FOUND, not a fixed number 10.
            # This prevents the IndexError crash.
            for _ in range(len(details)):
                # YOUR SCRIPT LOGIC IS PRESERVED EXACTLY:
                if _ % 2 != 0:
                    data.append(details[_].find_elements(By.TAG_NAME,"li")[-1].text)
                    
            # This update loop from your script is correct and is preserved.
            for k in range(len(data)):
                detail = data[k].split('\n')
                # Check to prevent error if split doesn't produce enough parts
                if len(detail) >= 3:
                    df.loc[k + i, 'Delivery Report'] = detail[0]
                    # Use partition for a safer split on the first space only
                    date_part, _, time_part = detail[1].partition(' ')
                    df.loc[k + i, 'date'] = date_part
                    df.loc[k + i, 'time'] = time_part
                    df.loc[k + i, 'office'] = detail[2]

            # The final increment is now simple and directly reflects the batch size.
            # It achieves the exact same result as your 'i += j; i += 1' logic.
            i += items_in_this_batch
            df_view.dataframe(df)

            elapsed = str(datetime.timedelta(seconds=int(time() - rt))).split(':')
            st.write(f"{i}) Record {ref} Completed - {elapsed[1]}:{elapsed[2]}")

            rt = 0

        # After loop
        ot = str(datetime.timedelta(seconds=int(time() - ot)))
        st.write(f"Total time :- {ot}")

    return df,pdfs


#For Barcode Generation

def generate_barcode_with_text(data, barcode_type='code128'):
    barcode_class = barcode.get_barcode_class(barcode_type)
    barcode_instance = barcode_class(data, writer=ImageWriter())

    # Save the barcode as an image file in memory without the default text
    options = {'write_text': False}
    barcode_image_io = BytesIO()
    barcode_instance.write(barcode_image_io, options=options)
    barcode_image_io.seek(0)
    
    # Open the barcode image
    custom_text = data[:2]+' '+data[2:-3]+' '+data[-3:]

    barcode_image = Image.open(barcode_image_io)

    # Get dimensions of the barcode image
    width, height = barcode_image.size

    # Create a blank image with the same width and additional height for text
    combined_image = Image.new('RGB', (width, height + 50), color='white')  # Increased height for text

    # Paste the barcode image onto the combined image
    combined_image.paste(barcode_image, (0, 0))

    # Get a drawing context
    draw = ImageDraw.Draw(combined_image)

    # Load a font
    try:
        # Load a TTF font file
        font = ImageFont.truetype("arial.ttf", 40)
    except IOError:
        # If the TTF font file is not found, use the default PIL font
        font = ImageFont.load_default()

    # Get text size using textbbox (text bounding box)
    text_bbox = draw.textbbox((0, 0), custom_text, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]

    # Calculate the position to draw the text to be centered below the barcode
    text_x = (width - text_width) // 2
    text_y = height + 10  # 10 pixels below the barcode image

    # Draw the text onto the combined image
    draw.text((text_x, text_y), custom_text, fill='black', font=font)

    # Save the combined image with text to memory
    combined_image_io = BytesIO()
    combined_image.save(combined_image_io, format='PNG')
    combined_image_io.seek(0)
    
    return combined_image_io

# Function to create a zip file with barcode images and return it in memory
def create_zip_with_barcodes(df,pth):
    zip_buffer = BytesIO()
    rns = df['RN']
    paths = []
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for rn in rns:
            barcode_image_io = generate_barcode_with_text(rn)
            zip_file.writestr(f"{rn}.png", barcode_image_io.read())
            paths.append(f"{pth}/barcodes/{rn}.png")
        df['code'] = paths
        # Write the DataFrame to an Excel file in memory
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        excel_buffer.seek(0)
        
        # Add the Excel file to the ZIP archive
        zip_file.writestr("updated_excel.xlsx", excel_buffer.read())
    zip_buffer.seek(0)
    return zip_buffer
    

import streamlit as st

if 'df' not in st.session_state:
    st.session_state.df = None
    st.session_state.cols = None

page = st.sidebar.radio("Select the Process", ["Status Extraction", "Hyperlink Assingment","Barcode Generation","PDF Name Changer"])

if page == "Status Extraction":
    uploaded_file = st.file_uploader("Upload the Excel file", type=["xlsx"])
    cols = st.columns(5)

    # Create a file uploader widget
    with cols[0]:
        start_ = st.text_input('Start at : ',placeholder='Index')
    with cols[1]:
        end = st.text_input('End at : ',placeholder='Index')
    with cols[2]:
        sleep_ = st.text_input('Limit : ',placeholder='Secounds')
    with cols[3]:
        pdf_opt = st.checkbox("Generate PDF's")
    with cols[4]:
        st.write()
        st.write()
        bt = st.button('START',help='Click to start the process')
    # Check if a file was uploaded
    if bt:
        if  uploaded_file is not None:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(uploaded_file)
            if len(list(df.columns)) != 7:
                st.error('ERROR!!! Invalid Excel Format')
            df.columns = ['Loan No','Name','RPAD Barcode No ','date','time','office','Delivery Report']
            print('1')
            for i in ['Name','RPAD Barcode No ','date','time','office','Delivery Report','Loan No']:
                df[i] = df[i].astype(str)

            if start_ == '' or not start_.isdigit():
                start_ = 1
            else:
                start_ = int(start_)
            if end == '' or not end.isdigit():
                end = len(df['RPAD Barcode No '])
            else:
                end = int(end)
            if sleep_ == '' or not sleep_.isdigit():
                sleep_ = 4
            else:
                sleep_ = int(sleep_)
            print('2')
            df,pdfs = start(df,start_,end,sleep_,pdf_opt)
            zip_data = BytesIO()
            with zipfile.ZipFile(zip_data, 'w') as zipf:
            # Add Excel file to the zip folder with a custom file name
                excel_file = BytesIO()
                with pd.ExcelWriter(excel_file, engine='xlsxwriter', mode='w') as writer:
                    df.to_excel(writer, index=False)
                excel_file.seek(0)
                zipf.writestr('output.xlsx', excel_file.read())
        
        # Provide Excel content as binary data to the download_button
    #         st.download_button(label="Download Excel", data=excel_content, file_name="output.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            if pdf_opt:
                # Create a zip file in memory
                with zipfile.ZipFile(zip_data, 'a') as zipf:
                    for pdf_data, pdf_name in pdfs:
                        zipf.writestr(pdf_name, b64decode(pdf_data))
                # Provide a download button for the zip file
            st.download_button(label='Download Files', data=zip_data.getvalue(), file_name='output.zip', mime='application/zip',help="Click to Download Excel File and PDF's")
            
        else:
            st.error('No file Selected!!!')
    #

elif page == "Hyperlink Assingment":
    cols = st.columns(4)
    dire = st.text_input("Select The Folder : ",placeholder='Enter the Path or select the directory')

    bt = st.button('Select Directory')
    if bt :
        path = askdirectory()
        excel_file_path = path+'/output.xlsx'
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        df = pd.read_excel(path+'/output.xlsx')
        data = {'Filename':[],'URL':[]}
        for i in list(df['Loan No']):
            data['URL'].append(str(i)+'.pdf')
            data['Filename'].append(str(i))
            
        # Example DataFrame with filenames and corresponding URLs
        df = pd.DataFrame(data)
        excel_file_path = path+'//output.xlsx'
        # Add headers
        
        
        for index, row in df.iterrows():
            filename = row['Filename']
            url = path+'/'+row['URL']
            
            # Find the cell corresponding to the filename in the first column
            for cell in worksheet['B']:
                if cell.value == filename:
                    # Create a hyperlink to the URL
                    cell.font = Font(underline="single", color="0000FF")
                    cell.hyperlink = url
                    break

      

        # Save the workbook to a file
        workbook.save(excel_file_path)

elif page == "Barcode Generation":
    st.title("Barcode Generator")

    # File uploader
    uploaded_file = st.file_uploader("Upload an Excel file containing RN numbers ", type=["xlsx"])

    st.write('Select Output Directory : ')

    if st.button('select'):
        if uploaded_file is None:
            st.error('Select Input File First')
        else:
            pth = askdirectory()        # Read the Excel file
            df = pd.read_excel(uploaded_file)
            try:
                df.columns = ['RN','code']
                zip_buffer = create_zip_with_barcodes(df,pth)

            # Provide download link
                st.download_button(
                    label="Download ZIP file with barcodes",
                    data=zip_buffer,
                    file_name="barcodes.zip",
                    mime="application/zip"
                )
            except Exception as e:
                print(e)
                st.error('Invalid File Format')
            # Check if the DataFrame has the necessary column
            # Create zip with barcodes
            

elif page == "PDF Name Changer":
    st.title("PDF Name Changer")

    # File uploader
    uploaded_file = st.file_uploader("Upload an Excel File", type=["xlsx"])

    # If file is uploaded, process it
    if uploaded_file is not None:
        # If the sheet is not already loaded, read the sheet
        
        df = pd.ExcelFile(uploaded_file)
        opts = df.sheet_names        
        sheet = st.selectbox("Choose the sheet:", opts, index=0)
        st.session_state.sheet = sheet  # Store the selected sheet in session_state

        if sheet:
            # Store the DataFrame and columns in session_state
            st.session_state.df = pd.read_excel(uploaded_file, sheet_name=sheet)
            st.session_state.cols = list(st.session_state.df.columns)

    # If the sheet and DataFrame are loaded, use the stored DataFrame and columns
            df = st.session_state.df
            cols = st.session_state.cols

            # Column selection (current name and new name columns)
            ip_col = st.selectbox("Choose current name field: ", cols, index=0)
            op_col = st.selectbox("Choose new name field: ", cols, index=1)

            # Store column selections in session_state to avoid refreshing
            st.session_state.ip_col = ip_col
            st.session_state.op_col = op_col
            if st.button('continue'):
                # Select PDF directory
                st.write('Select PDF Directory: ')
                path = askdirectory()

                if path:
                    st.session_state.path = path  # Store the path in session_state
                    if ip_col == op_col:
                        st.error('Select unique columns')
                    else:
                        try:
                            # Ensure the selected columns exist in the DataFrame
                            l = list(df[op_col])
                            t = list(df[ip_col])
                        except KeyError:
                            st.error("Invalid column Name")
                            

                        # Renaming files
                        for i in range(len(l)):
                            source = f'"{path}\\{t[i-1]}.pdf"'
                            destination = f'"{path}\\{l[i-1]}.pdf"'
                        
                            try:
                                os.system(f'move {source} {destination}')
                                st.write(f'{l[i-1]} completed')
                            except FileNotFoundError:
                                st.error(f"File {source} not found!")
                            except Exception as e:
                                print(e)
                                st.error(f"Error renaming {source} to {destination}: {str(e)}")

                        st.success("COMPLETED !!!!")

    else:
        st.warning("Please upload an Excel file to get started.")

                                        
                
    
